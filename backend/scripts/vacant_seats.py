"""
vacant_seats.py
----------------
Generic Maharashtra DTE Vacant Seats PDF parser.

Handles multiple common DTE vacant seat PDF formats:

FORMAT A (Engineering / B.Tech style) — category-wise seat matrix per college:
    Institute Code - Institute Name
    Course Name
    | Category | Total | Available |
    ...

FORMAT B (Nursing / GNM / MBA style) — row-per-college with category columns:
    Sr | Institute Code | Institute Name | Course | OPEN | OBC | SC | ST | EWS | NT | Total

FORMAT C (Simple text list):
    CollegeCode - CollegeName  CourseCode - CourseName  CategoryVacant...

The script auto-detects the format and falls back progressively.
"""

import pdfplumber
import pandas as pd
import re
import os


# ─────────────────────────────────────────────
# Category columns DTE commonly uses
# ─────────────────────────────────────────────
CATEGORY_COLS = [
    "GOPENS", "GSCS", "GSTS", "GOBCS", "GNTBS", "GNTCS", "GNTDS",
    "GVJAS", "GSEBCS", "GEWSS",
    "LOPENS", "LSCS", "LSTS", "LOBCS", "LNTBS", "LNTCS", "LNTDS",
    "LVJAS", "LSEBCS", "LEWSS",
    "OPEN", "SC", "ST", "OBC", "NTB", "NTC", "NTD", "VJA", "SEBC", "EWS",
    "TFWS", "PWD", "MI", "DEF",
    "G-OPEN", "G-SC", "G-ST", "G-OBC",
    "L-OPEN", "L-SC", "L-ST", "L-OBC",
    "TOTAL",
]

SKIP_LINES = [
    "GOVERNMENT OF MAHARASHTRA",
    "State Common Entrance",
    "Directorate of Technical",
    "DTE Maharashtra",
    "Printed On",
    "Page",
    "Note:",
    "Sr.No",
    "Sr. No",
    "---",
    "===",
    "Institute Code",
    "Choice Code",
    "Legends",
]


# ─────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────

def process(pdf_path, output_path, progress_callback=None):
    """
    Parse vacant seats PDF and write structured Excel output.
    Returns {"success": True/False, "records": N, "error": "..."}
    """
    def _cb(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        print(f"[{pct}%] {msg}", flush=True)

    try:
        _cb(5, "Opening Vacant Seats PDF...")

        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            _cb(8, f"PDF loaded — {total_pages} pages. Detecting format...")

            # Peek at first page to decide format
            first_text = pdf.pages[0].extract_text() or ""
            fmt = _detect_format(first_text)
            _cb(12, f"Detected format: {fmt}")

            if fmt == "TABLE":
                records = _parse_table_format(pdf, total_pages, _cb)
            elif fmt == "MATRIX":
                records = _parse_matrix_format(pdf, total_pages, _cb)
            else:
                records = _parse_text_format(pdf, total_pages, _cb)

        if not records:
            _cb(85, "Primary parse empty — trying generic fallback...")
            records = _generic_fallback(pdf_path, _cb)

        if not records:
            return {"success": False, "error": "No records could be extracted from this PDF. Please check the format."}

        _cb(88, f"Building Excel — {len(records)} rows...")
        df = _build_dataframe(records)
        _write_excel(df, output_path)

        _cb(100, f"Done! {len(df)} vacant seat records extracted.")
        return {"success": True, "records": len(df)}

    except Exception as e:
        return {"success": False, "error": str(e)}


# ─────────────────────────────────────────────
# Format detection
# ─────────────────────────────────────────────

def _detect_format(text):
    """
    Guess the PDF layout type from first page text.
    Returns 'TABLE', 'MATRIX', or 'TEXT'
    """
    upper = text.upper()
    # If it has pipe characters or grid-like structure → TABLE
    if text.count("|") > 10:
        return "TABLE"
    # If category headers appear in a row → MATRIX
    cat_found = sum(1 for c in ["OPEN", "OBC", "SC", "ST", "EWS"] if c in upper)
    if cat_found >= 3:
        return "MATRIX"
    return "TEXT"


# ─────────────────────────────────────────────
# Format A: Table (pipe-separated or pdfplumber table extraction)
# ─────────────────────────────────────────────

def _parse_table_format(pdf, total_pages, cb):
    records = []
    for page_num, page in enumerate(pdf.pages):
        tables = page.extract_tables()
        for table in tables:
            if not table or len(table) < 2:
                continue
            header = [str(c).strip().upper() if c else "" for c in table[0]]
            for row in table[1:]:
                if not row:
                    continue
                rec = _row_to_record(header, row)
                if rec:
                    records.append(rec)
        page.flush_cache()
        pct = 15 + int((page_num + 1) / total_pages * 65)
        cb(pct, f"Table parse — page {page_num + 1}/{total_pages}")
    return records


def _row_to_record(header, row):
    """Map a table row onto a standardised dict using header names."""
    d = {}
    for i, h in enumerate(header):
        val = str(row[i]).strip() if i < len(row) and row[i] else ""
        d[h] = val

    # Try to identify key fields by common header names
    inst_code = (d.get("INSTITUTE CODE") or d.get("INST CODE") or
                 d.get("CODE") or d.get("SR") or "").strip()
    inst_name = (d.get("INSTITUTE NAME") or d.get("COLLEGE NAME") or
                 d.get("COLLEGE") or d.get("NAME") or "").strip()
    course    = (d.get("COURSE") or d.get("COURSE NAME") or
                 d.get("BRANCH") or "").strip()

    if not inst_name and not course:
        return None

    rec = {
        "Institute Code": inst_code,
        "Institute Name": inst_name,
        "Course": course,
    }

    # Add any category columns found
    for cat in CATEGORY_COLS:
        val = d.get(cat, "")
        if val:
            rec[cat] = val

    # Total vacant
    total = (d.get("TOTAL") or d.get("TOTAL VACANT") or
             d.get("AVAILABLE") or d.get("VACANCY") or "").strip()
    rec["Total Vacant"] = total

    return rec if any(rec.values()) else None


# ─────────────────────────────────────────────
# Format B: Matrix (category headers in top row)
# ─────────────────────────────────────────────

def _parse_matrix_format(pdf, total_pages, cb):
    """
    Parse pages where category names appear as column headers
    and each row is an institute+course.
    """
    records = []
    current_institute = {"code": "", "name": ""}

    for page_num, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        lines = text.split("\n")

        for line in lines:
            line = line.strip()
            if not line or any(s in line for s in SKIP_LINES):
                continue

            # Institute header line: "01002 - Govt College of Engg, Amravati"
            inst_m = re.match(r"^(\d{4,6})\s*[-–]\s*(.+)", line)
            if inst_m:
                current_institute = {
                    "code": inst_m.group(1).strip(),
                    "name": inst_m.group(2).strip(),
                }
                continue

            # Data line — try to parse numbers as category vacancies
            rec = _parse_matrix_line(line, current_institute)
            if rec:
                records.append(rec)

        page.flush_cache()
        pct = 15 + int((page_num + 1) / total_pages * 65)
        cb(pct, f"Matrix parse — page {page_num + 1}/{total_pages}")

    return records


def _parse_matrix_line(line, institute):
    """
    Try to extract a data row: course name followed by seat counts.
    e.g. 'Civil Engineering   12  8  4  2  3  1   30'
    """
    # Find all numbers in the line
    numbers = re.findall(r"\b(\d+)\b", line)
    if len(numbers) < 2:
        return None

    # Text before first number is likely course name
    first_num_pos = re.search(r"\b\d+\b", line).start()
    course = line[:first_num_pos].strip()

    if not course or len(course) < 3:
        return None

    # Last number is usually total; rest are category-wise
    total = numbers[-1]
    cat_values = numbers[:-1]

    rec = {
        "Institute Code": institute.get("code", ""),
        "Institute Name": institute.get("name", ""),
        "Course": course,
        "Total Vacant": total,
    }

    # Map numbers to known category labels if count matches
    common_order = ["OPEN", "OBC", "SC", "ST", "EWS", "NTB", "NTC", "NTD", "VJA", "SEBC"]
    for i, val in enumerate(cat_values):
        label = common_order[i] if i < len(common_order) else f"CAT_{i+1}"
        rec[label] = val

    return rec


# ─────────────────────────────────────────────
# Format C: Plain text fallback
# ─────────────────────────────────────────────

def _parse_text_format(pdf, total_pages, cb):
    """
    Generic text parser — looks for institute + seat count patterns.
    """
    records = []
    current_institute = {"code": "", "name": ""}
    current_course = ""

    for page_num, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        lines = text.split("\n")

        for line in lines:
            line = line.strip()
            if not line or any(s in line for s in SKIP_LINES):
                continue

            # Institute line
            inst_m = re.match(r"^(\d{4,6})\s*[-–:]\s*(.+)", line)
            if inst_m:
                current_institute = {
                    "code": inst_m.group(1).strip(),
                    "name": inst_m.group(2).strip(),
                }
                current_course = ""
                continue

            # Course line (10-digit code dash name)
            course_m = re.match(r"(\d{10}[A-Za-z]?)\s*[-–]\s*(.+)", line)
            if course_m:
                current_course = course_m.group(2).strip()
                continue

            # Seat data line — has numbers
            numbers = re.findall(r"\b(\d+)\b", line)
            if numbers and current_institute["name"]:
                total = numbers[-1]
                rec = {
                    "Institute Code": current_institute["code"],
                    "Institute Name": current_institute["name"],
                    "Course": current_course or line[:40].strip(),
                    "Total Vacant": total,
                    "Raw Line": line,
                }
                # Try to pick out category values
                common_order = ["OPEN", "OBC", "SC", "ST", "EWS"]
                for i, val in enumerate(numbers[:-1]):
                    label = common_order[i] if i < len(common_order) else f"CAT_{i+1}"
                    rec[label] = val
                records.append(rec)

        page.flush_cache()
        pct = 15 + int((page_num + 1) / total_pages * 65)
        cb(pct, f"Text parse — page {page_num + 1}/{total_pages}")

    return records


# ─────────────────────────────────────────────
# Generic fallback — brute-force table extraction
# ─────────────────────────────────────────────

def _generic_fallback(pdf_path, cb):
    """Last-resort: try pdfplumber table extraction on every page."""
    records = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    for row in table:
                        if not row or not any(row):
                            continue
                        cells = [str(c).strip() if c else "" for c in row]
                        # Skip header rows
                        if cells[0].upper() in ("SR", "SR NO", "SR.NO", "INSTITUTE CODE", "CODE"):
                            continue
                        # At least one numeric cell = data row
                        if any(c.isdigit() for c in cells):
                            rec = {
                                "Institute Code": cells[0] if len(cells) > 0 else "",
                                "Institute Name": cells[1] if len(cells) > 1 else "",
                                "Course": cells[2] if len(cells) > 2 else "",
                            }
                            # Remaining cells as category data
                            for i, val in enumerate(cells[3:], start=1):
                                rec[f"Col_{i}"] = val
                            records.append(rec)
                page.flush_cache()
    except Exception as e:
        print(f"[Fallback error] {e}")
    return records


# ─────────────────────────────────────────────
# DataFrame builder
# ─────────────────────────────────────────────

def _build_dataframe(records):
    """Convert list of dicts to a clean DataFrame with consistent columns."""
    df = pd.DataFrame(records)

    # Ensure core columns always exist
    for col in ["Institute Code", "Institute Name", "Course", "Total Vacant"]:
        if col not in df.columns:
            df[col] = ""

    # Move core columns to front
    core = ["Institute Code", "Institute Name", "Course"]
    cat_cols_present = [c for c in CATEGORY_COLS if c in df.columns]
    other_cols = [c for c in df.columns if c not in core + cat_cols_present + ["Total Vacant", "Raw Line"]]

    ordered = core + cat_cols_present + ["Total Vacant"] + other_cols
    # Only keep columns that exist
    ordered = [c for c in ordered if c in df.columns]
    df = df[ordered]

    # Drop "Raw Line" if it crept in and is noisy
    if "Raw Line" in df.columns and len(df.columns) > 4:
        df.drop(columns=["Raw Line"], inplace=True)

    # Drop fully empty rows
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


# ─────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────

def _write_excel(df, output_path):
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vacant Seats")
        ws = writer.sheets["Vacant Seats"]

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Style header
        header_fill = PatternFill(start_color="1F5C2E", end_color="1F5C2E", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Freeze header row
        ws.freeze_panes = "A2"