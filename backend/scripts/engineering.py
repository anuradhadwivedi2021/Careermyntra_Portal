import pdfplumber
import re
import pandas as pd
import os
import gc

college_pattern      = re.compile(r"^\s*(\d{5})\s*-\s*(.+)", re.MULTILINE)
course_code_re       = re.compile(r"(\d{10}[A-Za-z]?)")
course_full_re       = re.compile(r"(\d{10}[A-Za-z]?)\s*-\s*(.+)")
status_pattern       = re.compile(r"Status:\s*(.+)")
home_university_pattern = re.compile(r"Home University\s*:\s*(.+)")

CATEGORY_KEYWORDS = [
    "OPEN", "OBC", "SC", "ST", "NTA", "NTB", "NTC", "NTD",
    "SEBC", "MI", "EWS", "TFWS", "ORPHAN", "VJ", "PWD"
]

COLUMNS = [
    "institute code", "institute name", "branch code", "course name",
    "status", "university", "category", "rank", "percentile",
    "gender", "quota", "category(1)", "branch", "year", "cap round"
]

def get_category1(category: str) -> str:
    if category.startswith("PWD"):
        return "PWD"
    for kw in CATEGORY_KEYWORDS:
        if kw in category:
            return kw
    return category[1:-1] if len(category) > 2 else category

def extract_rows(table, institute_code, university_name, branch_code, course_name, status, university):
    rows_data = []
    if len(table) < 2:
        return rows_data
    categories = [c.strip() if c else "" for c in table[0][1:]]
    for row in table[1:]:
        for col_idx, category in enumerate(categories):
            if not category or col_idx + 1 >= len(row):
                continue
            cell = row[col_idx + 1]
            if not cell or "(" not in cell:
                continue
            cell = cell.replace("\n", "")
            try:
                rank_str, rest = cell.split("(", 1)
                percentile_str = rest.rstrip(")")
                rank_str, percentile_str = rank_str.strip(), percentile_str.strip()
                if not rank_str:
                    continue
            except Exception:
                continue
            rows_data.append([
                institute_code, university_name, branch_code, course_name,
                status, university, category, rank_str, percentile_str,
                category[0] if category else "", category[-1] if category else "",
                get_category1(category), "B.Tech", "2025", "3"
            ])
    return rows_data

def process(pdf_path: str, output_path: str, progress_callback=None) -> dict:
    extracted_data = []
    def update(pct, msg):
        if progress_callback: progress_callback(pct, msg)
        print(f"[{pct}%] {msg}", flush=True)

    try:
        update(5, "Opening PDF...")
        
        # Use lower resolution for faster processing
        with pdfplumber.open(pdf_path, laparams={"line_margin": 0.5}) as pdf:
            total = len(pdf.pages)
            update(10, f"PDF loaded — {total} pages")
            
            for page_num, page in enumerate(pdf.pages):
                try:
                    # Extract text only once per page
                    text = page.extract_text()
                    if not text:
                        continue

                    college_match = college_pattern.search(text)
                    if not college_match:
                        continue

                    institute_code = int(college_match.group(1).strip())
                    university_name = college_match.group(2).strip()
                    
                    sm = status_pattern.search(text)
                    hm = home_university_pattern.search(text)
                    status   = sm.group(1).strip() if sm else "Unknown"
                    university = hm.group(1).strip() if hm else "Unknown"

                    # Extract course codes from text directly (faster than extract_words)
                    course_positions = []
                    for line in text.split("\n"):
                        m = course_code_re.search(line)
                        if m:
                            course_positions.append((len(course_positions), m.group(1).strip()))
                    
                    if not course_positions:
                        continue

                    course_name_map = {m.group(1).strip(): m.group(2).strip() for m in course_full_re.finditer(text)}
                    
                    # Use crop-based table extraction (faster than find_tables)
                    found_tables = page.find_tables(table_settings={
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 3,
                        "join_tolerance": 3,
                        "edge_min_length": 3,
                        "min_words_vertical": 1,
                        "min_words_horizontal": 1,
                    })

                    for ft in found_tables:
                        table_data = ft.extract()
                        if not table_data or len(table_data) < 2:
                            continue
                        # Assign course based on position
                        assigned_course = course_positions[-1][1] if course_positions else ""
                        cname = course_name_map.get(assigned_course, assigned_course)
                        extracted_data.extend(extract_rows(
                            table_data, institute_code, university_name,
                            assigned_course, cname, status, university
                        ))

                except Exception as page_err:
                    print(f"[WARN] Page {page_num+1} error: {page_err}", flush=True)
                finally:
                    page.flush_cache()

                # Progress update every 50 pages + gc every 100
                if (page_num + 1) % 50 == 0 or page_num == total - 1:
                    pct = 10 + int(((page_num + 1) / total) * 75)
                    update(pct, f"Processing page {page_num+1}/{total} — {len(extracted_data)} records")
                if (page_num + 1) % 100 == 0:
                    gc.collect()

        gc.collect()
        update(88, f"Building Excel — {len(extracted_data)} records...")
        
        df = pd.DataFrame(extracted_data, columns=COLUMNS)
        df["institute code"] = pd.to_numeric(df["institute code"], errors="coerce").astype("Int64")
        df["rank"]           = pd.to_numeric(df["rank"], errors="coerce")
        df["percentile"]     = pd.to_numeric(df["percentile"], errors="coerce")
        df["cap round"]      = pd.to_numeric(df["cap round"], errors="coerce").astype("Int64")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Cut-off Data")
            ws = writer.sheets["Cut-off Data"]

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill("solid", fgColor="1565C0")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin        = Side(style="thin", color="D0D7E3")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            center      = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            col_widths = {
                "institute code": 14, "institute name": 40,
                "branch code": 14, "course name": 35,
                "status": 22, "university": 28
            }
            for i, col in enumerate(COLUMNS, 1):
                ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 12)

            light = PatternFill("solid", fgColor="EFF4FF")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                is_even = row[0].row % 2 == 0
                for cell in row:
                    if is_even:
                        cell.fill = light
                    cell.alignment = center
                    cell.border = border

            ws.freeze_panes      = "A2"
            ws.auto_filter.ref   = ws.dimensions

        update(100, "Done!")
        return {"success": True, "records": len(df), "output_path": output_path, "error": None}

    except Exception as e:
        import traceback
        print(f"[ERROR] {e}\n{traceback.format_exc()}", flush=True)
        return {"success": False, "records": 0, "output_path": None, "error": str(e)}