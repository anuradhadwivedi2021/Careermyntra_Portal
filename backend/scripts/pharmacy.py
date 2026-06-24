# pharmacy.py — CareerMyntra Backend Script
# Processes: B.Pharm / Pharm.D Cut-offs CAP Round PDF
# Output: Structured Excel with 15 columns (same format as engineering.py)

import pdfplumber
import re
import pandas as pd
import os

# ─── Patterns ────────────────────────────────────────────────
college_pattern  = re.compile(r"^\s*(\d{5})\s*-\s*(.+)", re.MULTILINE)
course_code_re   = re.compile(r"(\d{10}[A-Za-z]?)")
course_full_re   = re.compile(r"(\d{10}[A-Za-z]?)\s*-\s*(.+)")
status_pattern   = re.compile(r"Status:\s*(.+)")
# Pharmacy PDFs don't have "Home University :" line — use "Unknown" fallback
home_univ_pattern = re.compile(r"Home University\s*:\s*(.+)")

CATEGORY_KEYWORDS = [
    "OPEN", "OBC", "SC", "ST", "NTA", "NTB", "NTC", "NTD",
    "SEBC", "MI", "EWS", "TFWS", "ORPHAN", "VJ", "PWD"
]

COLUMNS = [
    "institute code", "institute name", "branch code", "course name",
    "status", "university", "category", "rank", "percentile",
    "gender", "quota", "category(1)", "branch", "year", "cap round"
]

# ─── Helpers ─────────────────────────────────────────────────
def get_category1(category: str) -> str:
    """Map raw category code to simplified category name."""
    if category.startswith("PWD"):
        return "PWD"
    for kw in CATEGORY_KEYWORDS:
        if kw in category:
            return kw
    return category[1:-1] if len(category) > 2 else category


def detect_cap_round(pdf_path: str, first_page_text: str = "") -> str:
    """Detect CAP round number from filename or PDF content."""
    name = os.path.basename(pdf_path).lower()
    text = (name + " " + first_page_text[:300]).lower()

    roman = {"iii": "3", "ii": "2", "i": "1", "iv": "4", "v": "5"}
    for r, n in roman.items():
        if re.search(rf"cap[- _]*round[- _]*{r}\b", text):
            return n
        if re.search(rf"cap[- _]*{r}\b", text):
            return n
    m = re.search(r"cap\s*(\d)", text)
    if m:
        return m.group(1)
    m = re.search(r"round[- _]*(\d)", text)
    if m:
        return m.group(1)
    return "1"


def detect_year(pdf_path: str, first_page_text: str = "") -> str:
    """Detect academic year from filename or PDF content."""
    text = os.path.basename(pdf_path) + " " + first_page_text[:300]
    m = re.search(r"20(\d{2})", text)
    return "20" + m.group(1) if m else "2025"


def extract_rows(table, institute_code, university_name,
                 branch_code, course_name, status, university,
                 cap_round, year):
    """Extract data rows from a pdfplumber table."""
    rows_data = []
    if len(table) < 2:
        return rows_data

    categories = [c.strip() if c else "" for c in table[0][1:]]

    for row in table[1:]:
        for col_idx, category in enumerate(categories):
            if not category:
                continue
            cell_idx = col_idx + 1
            if cell_idx >= len(row):
                continue
            cell = row[cell_idx]
            if not cell or "(" not in cell:
                continue

            cell = cell.replace("\n", "")
            try:
                rank_str, rest = cell.split("(", 1)
                percentile_str = rest.rstrip(")")
                rank_str       = rank_str.strip()
                percentile_str = percentile_str.strip()
                if not rank_str:
                    continue
            except Exception:
                continue

            gender = category[0] if category else ""
            quota  = category[-1] if category else ""
            cat1   = get_category1(category)

            rows_data.append([
                institute_code, university_name, branch_code, course_name,
                status, university, category, rank_str, percentile_str,
                gender, quota, cat1, "B.Pharm", year, cap_round
            ])

    return rows_data


# ─── Main Process Function ───────────────────────────────────
def process(pdf_path: str, output_path: str, progress_callback=None) -> dict:
    """
    Main function called by Flask backend.

    Args:
        pdf_path:          Path to uploaded PDF file
        output_path:       Path where output Excel should be saved
        progress_callback: Optional function(percent, message) for live updates

    Returns:
        dict with keys: success, records, output_path, error
    """
    import gc

    extracted_data = []

    def update(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        print(f"[{pct}%] {msg}")

    try:
        update(5, "Opening PDF...")

        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            update(10, f"PDF loaded — {total} pages found")

            # ── Auto-detect round & year from filename + first page ──
            first_text = pdf.pages[0].extract_text() or "" if pdf.pages else ""
            cap_round  = detect_cap_round(pdf_path, first_text)
            year       = detect_year(pdf_path, first_text)
            update(12, f"Detected: B.Pharm | CAP Round {cap_round} | Year {year}")

            for page_num, page in enumerate(pdf.pages):
                # ── Extract text ──
                text = page.extract_text()
                if not text:
                    # FIX: release this page's cached objects even when skipped
                    page.flush_cache()
                    continue

                # ── College info ──
                college_match = college_pattern.search(text)
                if not college_match:
                    page.flush_cache()
                    continue

                institute_code  = int(college_match.group(1).strip())
                university_name = college_match.group(2).strip()

                status_match = status_pattern.search(text)
                home_match   = home_univ_pattern.search(text)

                status     = status_match.group(1).strip() if status_match else "Unknown"
                # Pharmacy PDFs have no Home University field → use "Maharashtra"
                university = home_match.group(1).strip() if home_match else "Maharashtra"

                # ── Course codes from word positions ──
                words = page.extract_words()
                course_positions = [
                    (w["top"], w["text"].strip())
                    for w in words
                    if course_code_re.fullmatch(w["text"].strip())
                ]
                course_positions.sort(key=lambda x: x[0])

                if not course_positions:
                    page.flush_cache()
                    continue

                # ── Course name map ──
                course_name_map = {
                    m.group(1).strip(): m.group(2).strip()
                    for m in course_full_re.finditer(text)
                }

                # ── Tables ──
                found_tables = page.find_tables()
                for ft in found_tables:
                    table_top  = ft.bbox[1]
                    table_data = ft.extract()
                    if not table_data or len(table_data) < 2:
                        continue

                    # Assign nearest course code above this table
                    assigned_course = None
                    best_y = -1
                    for (cy, code) in course_positions:
                        if cy <= table_top and cy > best_y:
                            best_y = cy
                            assigned_course = code

                    if not assigned_course:
                        assigned_course = course_positions[0][1]

                    cname = course_name_map.get(assigned_course, assigned_course)
                    rows  = extract_rows(
                        table_data, institute_code, university_name,
                        assigned_course, cname, status, university,
                        cap_round, year
                    )
                    extracted_data.extend(rows)

                # FIX: pdfplumber caches each page's parsed objects (chars, lines,
                # rects, etc.) internally — for a 500+ page PDF this adds up to
                # gigabytes of RAM. flush_cache() releases that page's cache
                # once we're done reading it, since we never need to re-read it.
                page.flush_cache()

                # FIX: every 25 pages, force Python's garbage collector to run.
                # pdfplumber/pdfminer create a lot of short-lived objects per
                # page; without an explicit gc.collect(), memory can stay
                # fragmented and "used" even after objects are unreferenced.
                if (page_num + 1) % 25 == 0:
                    gc.collect()

                # ── Progress update ──
                pct = 10 + int(((page_num + 1) / total) * 75)
                if (page_num + 1) % 50 == 0 or page_num == total - 1:
                    update(pct, f"Processing page {page_num+1}/{total} — {len(extracted_data)} records")

        # FIX: PDF is now closed — drop any remaining references and collect
        # before the memory-heavy DataFrame/Excel-writing stage begins.
        gc.collect()

        update(88, "Building Excel file...")

        # ── Build DataFrame ──
        df = pd.DataFrame(extracted_data, columns=COLUMNS)
        df["institute code"] = pd.to_numeric(df["institute code"], errors="coerce").astype("Int64")
        df["rank"]           = pd.to_numeric(df["rank"],           errors="coerce")
        df["percentile"]     = pd.to_numeric(df["percentile"],     errors="coerce")
        df["cap round"]      = pd.to_numeric(df["cap round"],      errors="coerce").astype("Int64")

        update(93, "Applying Excel formatting...")

        # ── Write Excel with formatting ──
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Cut-off Data")

            wb = writer.book
            ws = writer.sheets["Cut-off Data"]

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Header style
            header_fill = PatternFill("solid", fgColor="1565C0")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin   = Side(style="thin", color="D0D7E3")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border

            ws.row_dimensions[1].height = 30

            # Column widths
            col_widths = {
                "institute code": 14, "institute name": 40, "branch code": 14,
                "course name":    35, "status":         22, "university":  28,
                "category":       12, "rank":           10, "percentile":  14,
                "gender":          8, "quota":           8, "category(1)": 14,
                "branch":         10, "year":            8, "cap round":   10
            }
            for i, col in enumerate(COLUMNS, 1):
                ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 14)

            # Alternate row shading
            light = PatternFill("solid", fgColor="EFF4FF")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.row % 2 == 0:
                        cell.fill = light
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

            # Freeze header row + auto-filter
            ws.freeze_panes        = "A2"
            ws.auto_filter.ref     = ws.dimensions

        update(100, f"Done! {len(df)} records saved.")

        return {
            "success":     True,
            "records":     len(df),
            "output_path": output_path,
            "error":       None
        }

    except Exception as e:
        update(0, f"Error: {str(e)}")
        return {
            "success":     False,
            "records":     0,
            "output_path": None,
            "error":       str(e)
        }


# ─── Standalone Run ──────────────────────────────────────────
if __name__ == "__main__":
    import sys
    pdf = sys.argv[1] if len(sys.argv) > 1 else r"Cap 1 Pharma 25-26.pdf"
    out = sys.argv[2] if len(sys.argv) > 2 else "cutoff_BPharm_cap1_output.xlsx"
    result = process(pdf, out)
    print(result)